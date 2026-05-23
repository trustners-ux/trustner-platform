"""Convert TRUSTNER FUND LIST Excel to TS category files matching the existing schema.

Also resolves AMFI scheme codes for every fund so the live-data fund detail pages
work end-to-end. Strategy:
  1. Re-use schemeCodes from any prior month's data files (fast, free, exact match).
  2. Fall back to MFAPI.in search for unmatched funds (free, ~1s/fund).

After running this script you should also run ``patch_scheme_codes.py`` to backfill
any missing schemeCodes — but for newly written files this is integrated by default.
"""
import json
import os
import re
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Optional, Dict
from openpyxl import load_workbook

TRUSTNER_DIR = Path("/Users/ram/Documents/Trustner Tech Project/Mera SIPOnline/sip-platform/src/data/funds/trustner")


def _norm_name(name: str) -> str:
    s = name.lower()
    s = re.sub(r"\(.*?\)", "", s)
    s = re.sub(r"[-–—]", " ", s)
    s = re.sub(r"\b(regular|direct|plan|growth|idcw|fund|the|of)\b", " ", s)
    s = re.sub(r"[^a-z0-9 ]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _build_scheme_code_lookup() -> Dict[str, int]:
    """Walk every prior month's TS file and harvest name→schemeCode pairs."""
    lookup: Dict[str, int] = {}
    name_re = re.compile(r"name:\s*'([^']+)'")
    code_re = re.compile(r"schemeCode:\s*(\d+)")
    for month_dir in TRUSTNER_DIR.iterdir():
        if not month_dir.is_dir():
            continue
        for fp in month_dir.glob("*.ts"):
            if fp.name == "index.ts":
                continue
            text = fp.read_text()
            for block in re.split(r"^\s*\{\s*$", text, flags=re.MULTILINE):
                nm = name_re.search(block)
                cd = code_re.search(block)
                if nm and cd:
                    lookup[_norm_name(nm.group(1))] = int(cd.group(1))
    return lookup


def _mfapi_lookup(name: str) -> Optional[int]:
    """Search MFAPI.in for the closest matching Regular Plan Growth scheme."""
    try:
        q = re.sub(r"\(.*?\)", "", name).strip()
        url = f"https://api.mfapi.in/mf/search?q={urllib.parse.quote(q)}"
        req = urllib.request.Request(url, headers={"User-Agent": "Trustner-Fund-Importer"})
        with urllib.request.urlopen(req, timeout=8) as resp:
            data = json.loads(resp.read())
        if not data:
            return None
        target = set(_norm_name(name).split())
        scored = []
        for item in data:
            sn = item.get("schemeName", "")
            sc = item.get("schemeCode")
            if not sc:
                continue
            score = 0
            if "regular" in sn.lower(): score += 10
            if "direct" in sn.lower(): score -= 15
            if "growth" in sn.lower(): score += 5
            if "idcw" in sn.lower() or "dividend" in sn.lower(): score -= 8
            score += len(target & set(_norm_name(sn).split())) * 2
            scored.append((score, sc))
        scored.sort(reverse=True)
        return int(scored[0][1]) if scored and scored[0][0] > 5 else None
    except Exception:
        return None


def resolve_scheme_code(fund_name: str, lookup: Dict[str, int]) -> Optional[int]:
    """Try prior-month lookup first, then MFAPI fallback."""
    norm = _norm_name(fund_name)
    if norm in lookup:
        return lookup[norm]
    target_tokens = set(norm.split())
    for k, v in lookup.items():
        if target_tokens and target_tokens.issubset(set(k.split())):
            return v
    return _mfapi_lookup(fund_name)

EXCEL = "/Users/ram/Documents/Trustner Tech Project/Mera SIPOnline/Fund List/TRUSTNER FUND LIST FOR THE MONTH OF MAY'26.xlsx"
OUT = "/Users/ram/Documents/Trustner Tech Project/Mera SIPOnline/sip-platform/src/data/funds/trustner/may-2026"
MONTH = "May"
YEAR = 2026
DATA_AS_ON = "01.05.2026"
LAST_UPDATED = "2026-05-01"

# Section name (Excel) -> (FundCategory.name, displayName, hasSkinInTheGame, varName, fileName)
CATEGORY_MAP = {
    "VALUE FUND":               ("Value",                "Value Funds",                True,  "VALUE_FUNDS",                "value-funds"),
    "SMALL CAP FUND":           ("Small Cap",            "Small Cap Funds",            True,  "SMALL_CAP_FUNDS",            "small-cap"),
    "MID CAP FUND":             ("Mid Cap",              "Mid Cap Funds",              True,  "MID_CAP_FUNDS",              "mid-cap"),
    "LARGE CAP FUND":           ("Large Cap",            "Large Cap Funds",            True,  "LARGE_CAP_FUNDS",            "large-cap"),
    "LARGE AND MIDCAP":         ("Large & Mid Cap",      "Large & Mid Cap Funds",      True,  "LARGE_MID_CAP_FUNDS",        "large-mid-cap"),
    "MULTI CAP FUND":           ("Multi Cap",            "Multi Cap Funds",            True,  "MULTI_CAP_FUNDS",            "multi-cap"),
    "MULTI ASSET FUND":         ("Multi Asset",          "Multi Asset Funds",          True,  "MULTI_ASSET_FUNDS",          "multi-asset"),
    "BALANCE ADVANTAGE FUND":   ("Balanced Advantage",   "Balanced Advantage Funds",   True,  "BALANCED_ADVANTAGE_FUNDS",   "balanced-advantage"),
    "FLEXI - CAP":              ("Flexi Cap",            "Flexi Cap Funds",            True,  "FLEXI_CAP_FUNDS",            "flexi-cap"),
    "AGGRESSIVE HYBRID FUND":   ("Aggressive Hybrid",    "Aggressive Hybrid Funds",    True,  "AGGRESSIVE_HYBRID_FUNDS",    "aggressive-hybrid"),
    "EQUITY SAVINGS FUND":      ("Equity Savings",       "Equity Savings Funds",       True,  "EQUITY_SAVINGS_FUNDS",       "equity-savings"),
    "CONSERVATIVE HYBRID FUND": ("Conservative Hybrid",  "Conservative Hybrid Funds",  True,  "CONSERVATIVE_HYBRID_FUNDS",  "conservative-hybrid"),
    "GOLD AND SILVER":          ("Gold & Silver",        "Gold & Silver Funds",        False, "GOLD_SILVER_FUNDS",          "gold-silver"),
}


def normalize_section(s: str) -> str:
    """Uppercase, collapse whitespace, drop punctuation noise."""
    s = re.sub(r"\s+", " ", s.strip().upper())
    return s


def slugify(name: str) -> str:
    s = name.lower()
    s = re.sub(r"[^a-z0-9\s-]", "", s)
    s = re.sub(r"\s+", "-", s)
    s = re.sub(r"-+", "-", s)
    return s.strip("-")


def is_header_row(cells) -> bool:
    """Detect header row — must contain 'Fund Name' or 'Fund Manager' near column 1-3."""
    joined = " | ".join(str(c).lower() for c in cells if c is not None)
    return "fund name" in joined and "fund manager" in joined


def is_section_header(cells) -> tuple:
    """If exactly one non-empty cell that matches a known category, return (section_name, normalized).
    Otherwise return (None, None)."""
    non_empty = [(i, c) for i, c in enumerate(cells) if c is not None and str(c).strip()]
    if len(non_empty) != 1:
        return None, None
    text = str(non_empty[0][1]).strip()
    norm = normalize_section(text)
    # Strip trailing punctuation/dashes
    for known in CATEGORY_MAP.keys():
        if normalize_section(known) == norm:
            return text, known
    return None, None


def find_col_idx(headers, *candidates) -> int:
    """Return column index whose lowercased header matches any candidate as a whole word/token.
    Uses word-boundary regex so 'age' won't false-match 'manager', 'sd' won't match 'holdings', etc."""
    for i, h in enumerate(headers):
        if h is None:
            continue
        s = str(h).lower().strip()
        for cand in candidates:
            # Match candidate as a whole token (handles "age (yrs)", "1y", "st. dev.", "% of aum")
            if re.search(rf"(?:^|[^a-z0-9]){re.escape(cand)}(?:[^a-z0-9]|$)", s):
                return i
    return -1


def to_float(v, default=0.0) -> float:
    if v is None:
        return default
    s = str(v).strip()
    if s in ("", "-", "NA", "N/A"):
        return default
    s = s.replace(",", "").replace("%", "")
    try:
        return float(s)
    except ValueError:
        return default


def to_int(v, default=0) -> int:
    return int(round(to_float(v, default)))


def escape_ts_string(s: str) -> str:
    return s.replace("\\", "\\\\").replace("'", "\\'")


def fund_to_ts(fund: dict) -> str:
    """Format a fund object into a TS literal matching the existing schema."""
    lines = [
        "    {",
        f"      id: '{fund['id']}',",
        f"      name: '{escape_ts_string(fund['name'])}',",
        f"      category: '{fund['category']}',",
        f"      fundManager: '{escape_ts_string(fund['fundManager'])}',",
        f"      ageOfFund: {fund['ageOfFund']},",
        f"      aumCr: {fund['aumCr']},",
        f"      ter: {fund['ter']},",
        f"      standardDeviation: {fund['standardDeviation']},",
        f"      sharpeRatio: {fund['sharpeRatio']},",
        "      returns: {",
        f"        mtd: {fund['returns']['mtd']},",
        f"        ytd: {fund['returns']['ytd']},",
        f"        oneYear: {fund['returns']['oneYear']},",
        f"        twoYear: {fund['returns']['twoYear']},",
        f"        threeYear: {fund['returns']['threeYear']},",
        f"        fiveYear: {fund['returns']['fiveYear']},",
        "      },",
        f"      numberOfHoldings: {fund['numberOfHoldings']},",
    ]
    if fund.get("skinInTheGame"):
        sig = fund["skinInTheGame"]
        lines.append("      skinInTheGame: {")
        lines.append(f"        amountCr: {sig['amountCr']},")
        lines.append(f"        percentOfAum: {sig['percentOfAum']},")
        lines.append("      },")
    lines.append(f"      rank: {fund['rank']},")
    if fund.get("schemeCode"):
        lines.append(f"      schemeCode: {fund['schemeCode']},")
    lines.append("    },")
    return "\n".join(lines)


def category_to_ts(cat: dict, var_name: str) -> str:
    funds_sorted = sorted(cat["funds"], key=lambda f: f["rank"] if f["rank"] > 0 else 999)
    fund_blocks = "\n".join(fund_to_ts(f) for f in funds_sorted)
    return f"""import {{ TrustnerFundCategory }} from '@/types/funds';

export const {var_name}: TrustnerFundCategory = {{
  name: '{cat['name']}',
  displayName: '{cat['displayName']}',
  hasSkinInTheGame: {str(cat['hasSkinInTheGame']).lower()},
  funds: [
{fund_blocks}
  ],
}};
"""


def main():
    wb = load_workbook(EXCEL, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    # Build the prior-month schemeCode lookup once. Funds that match by name
    # will inherit their scheme codes — no MFAPI calls needed for repeats.
    print("Building schemeCode lookup from prior months…")
    code_lookup = _build_scheme_code_lookup()
    print(f"  → {len(code_lookup)} prior-month name→code pairs available")

    sections = {}  # section_name -> list of fund dicts
    current_section = None
    current_headers = None

    for row in rows:
        cells = list(row)

        sec_text, sec_key = is_section_header(cells)
        if sec_key:
            current_section = sec_key
            current_headers = None
            sections[current_section] = []
            continue

        if current_section is None:
            continue

        if is_header_row(cells):
            current_headers = [str(c).strip().lower() if c is not None else "" for c in cells]
            continue

        if current_headers is None:
            continue

        # Data row — must have a fund name in the column where headers detected "fund name"
        name_col = find_col_idx(current_headers, "fund name")
        if name_col < 0:
            continue
        if name_col >= len(cells) or cells[name_col] is None:
            continue
        fund_name = str(cells[name_col]).strip()
        if not fund_name or fund_name.lower() == "fund name":
            continue

        # Extract values by header lookup
        manager_col = find_col_idx(current_headers, "fund manager")
        age_col = find_col_idx(current_headers, "age")
        aum_col = find_col_idx(current_headers, "aum")
        ter_col = find_col_idx(current_headers, "ter")
        sd_col = find_col_idx(current_headers, "st. dev", "std dev", "standard deviation", "sd")
        sharpe_col = find_col_idx(current_headers, "sharpe")
        mtd_col = find_col_idx(current_headers, "mtd")
        ytd_col = find_col_idx(current_headers, "ytd")
        y1_col = find_col_idx(current_headers, "1y")
        y2_col = find_col_idx(current_headers, "2y")
        y3_col = find_col_idx(current_headers, "3y")
        y5_col = find_col_idx(current_headers, "5y")
        holdings_col = find_col_idx(current_headers, "holdings")
        skin_pct_col = find_col_idx(current_headers, "skin in game", "% of aum")
        rank_col = find_col_idx(current_headers, "ranking")

        def cell(idx):
            if idx < 0 or idx >= len(cells):
                return None
            return cells[idx]

        cat_info = CATEGORY_MAP[current_section]
        cat_name = cat_info[0]

        fund_id = slugify(fund_name)
        manager = str(cell(manager_col)).strip() if cell(manager_col) else "—"

        aum_cr = to_int(cell(aum_col))
        skin_pct = to_float(cell(skin_pct_col))
        skin_amount = round(aum_cr * skin_pct, 4) if skin_pct > 0 else 0
        rank = to_int(cell(rank_col))

        scheme_code = resolve_scheme_code(fund_name, code_lookup)

        fund = {
            "id": fund_id,
            "name": fund_name,
            "category": cat_name,
            "fundManager": manager,
            "ageOfFund": to_float(cell(age_col)),
            "aumCr": aum_cr,
            "ter": to_float(cell(ter_col)),
            "standardDeviation": to_float(cell(sd_col)),
            "sharpeRatio": to_float(cell(sharpe_col)),
            "returns": {
                "mtd": to_float(cell(mtd_col)),
                "ytd": to_float(cell(ytd_col)),
                "oneYear": to_float(cell(y1_col)),
                "twoYear": to_float(cell(y2_col)),
                "threeYear": to_float(cell(y3_col)),
                "fiveYear": to_float(cell(y5_col)),
            },
            "numberOfHoldings": to_int(cell(holdings_col)),
            "rank": rank if rank > 0 else 99,
            "schemeCode": scheme_code,
        }
        if cat_info[2] and skin_pct > 0:  # hasSkinInTheGame
            fund["skinInTheGame"] = {
                "amountCr": skin_amount,
                "percentOfAum": skin_pct,
            }
        sections[current_section].append(fund)

    # ── Write category files ──
    os.makedirs(OUT, exist_ok=True)
    written = []
    for section, funds in sections.items():
        if not funds:
            print(f"⚠️  Skipping empty section: {section}")
            continue
        cat_name, display, has_skin, var_name, file_slug = CATEGORY_MAP[section]
        cat_obj = {
            "name": cat_name,
            "displayName": display,
            "hasSkinInTheGame": has_skin,
            "funds": funds,
        }
        ts = category_to_ts(cat_obj, var_name)
        fp = os.path.join(OUT, f"{file_slug}.ts")
        with open(fp, "w") as f:
            f.write(ts)
        print(f"✅ {file_slug}.ts — {len(funds)} funds")
        written.append((section, var_name, file_slug, cat_name))

    # ── Write index.ts ──
    # Maintain the same order the existing march-2026 index uses
    canonical_order = [
        "LARGE CAP FUND", "LARGE AND MIDCAP", "MID CAP FUND", "SMALL CAP FUND",
        "FLEXI - CAP", "MULTI CAP FUND", "VALUE FUND",
        "MULTI ASSET FUND", "BALANCE ADVANTAGE FUND",
        "AGGRESSIVE HYBRID FUND", "EQUITY SAVINGS FUND", "CONSERVATIVE HYBRID FUND",
        "GOLD AND SILVER",
    ]
    written_set = {w[0]: w for w in written}
    ordered = [written_set[s] for s in canonical_order if s in written_set]

    imports = "\n".join(f"import {{ {var_name} }} from './{file_slug}';" for _, var_name, file_slug, _ in ordered)
    cat_list = "\n    ".join(f"{var_name}," for _, var_name, _, _ in ordered)

    index_ts = f"""import {{ TrustnerFundList }} from '@/types/funds';
{imports}

export const {MONTH.upper()}_{YEAR}_FUND_LIST: TrustnerFundList = {{
  month: '{MONTH}',
  year: {YEAR},
  dataAsOn: '{DATA_AS_ON}',
  lastUpdated: '{LAST_UPDATED}',
  categories: [
    {cat_list}
  ],
}};
"""
    with open(os.path.join(OUT, "index.ts"), "w") as f:
        f.write(index_ts)
    print(f"✅ index.ts — {len(ordered)} categories")

    # ── Update parent trustner/index.ts (idempotent) ──
    parent_index_path = "/Users/ram/Documents/Trustner Tech Project/Mera SIPOnline/sip-platform/src/data/funds/trustner/index.ts"
    with open(parent_index_path) as f:
        parent = f.read()
    target_import = f"import {{ {MONTH.upper()}_{YEAR}_FUND_LIST }} from './{MONTH.lower()}-{YEAR}';"
    target_const = f"export const CURRENT_TRUSTNER_LIST: TrustnerFundList = {MONTH.upper()}_{YEAR}_FUND_LIST;"
    # Only add the import if it isn't already present
    if target_import not in parent:
        parent = re.sub(
            r"^(import \{[^}]*\} from '\./(?:march|april|may|june|july|august|september|october|november|december)-\d{4}';\s*\n)",
            r"\1" + target_import + "\n",
            parent,
            count=1,
            flags=re.MULTILINE,
        )
    # Replace the CURRENT_TRUSTNER_LIST assignment regardless of which month it was
    parent = re.sub(
        r"export const CURRENT_TRUSTNER_LIST: TrustnerFundList = [A-Z_]+;",
        target_const,
        parent,
    )
    with open(parent_index_path, "w") as f:
        f.write(parent)
    print(f"✅ Updated parent trustner/index.ts → CURRENT_TRUSTNER_LIST = {MONTH.upper()}_{YEAR}_FUND_LIST")

    total_funds = sum(len(s) for s in sections.values())
    print(f"\n🎉 Done. {total_funds} funds across {len(ordered)} categories written to may-2026/")


if __name__ == "__main__":
    main()
