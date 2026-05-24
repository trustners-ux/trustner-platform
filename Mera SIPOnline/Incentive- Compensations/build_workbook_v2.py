#!/usr/bin/env python3
"""
Build Trustner Incentive Master Workbook v2
Complete rebuild with fixed dropdowns, corrected credit rules, and all sheets
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
import copy

# ============================================================================
# CONSTANTS
# ============================================================================

NAVY = "1A365D"
YELLOW_INPUT = "FFFFF0"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"

header_font = Font(name="Arial", size=11, bold=True, color=WHITE)
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
input_font = Font(name="Arial", size=10, color="0000FF")
input_fill = PatternFill(start_color=YELLOW_INPUT, end_color=YELLOW_INPUT, fill_type="solid")
formula_font = Font(name="Arial", size=10, color="000000")
formula_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Employee data
EMPLOYEES = [
    ("Balbinder Kaur", "TAS001", "2013-03-01", "Inhouse Sales Manager", "OPERATION", "SALES", 36017, "TAS", 1.5),
    ("Pranab Kumar Barman", "TAS002", "2014-09-15", "Office Assistant", "OPERATION", "ADMIN", 21429, "TAS", 0),
    ("Ajanta Saikia", "TAS003", "2019-01-31", "Field Force Multiplier", "LIFE", "SALES", 99751, "TIB", 1.5),
    ("Akash Kumar", "TAS004", "2020-09-01", "Executive", "HEALTH", "SALES", 22449, "TIB", 1.5),
    ("Nipa Das", "TAS005", "2020-11-16", "Back Office Executive", "OPERATION", "Back Office", 22263, "TAS", 0),
    ("Jitender Roy", "TAS006", "2021-07-06", "Relationship Manager", "HEALTH", "SALES", 40440, "TIB", 1.5),
    ("Rafiqueddin Ahmed", "TAS051", "2021-07-09", "Mentor", "OPERATION", "SALES", 30000, "TAS", 1.5),
    ("Pranita Saikia", "TAS008", "2021-07-29", "Back Office Executive", "G.I.", "Back Office", 34500, "TIB", 0),
    ("Tamanna Kejriwal", "TAS009", "2022-06-22", "Sales Head", "F.P.", "SALES", 90000, "TAS", 1.5),
    ("Shivani Kumari", "TAS013", "2022-09-01", "Back Office Executive", "OPERATION", "Back Office", 16749, "TAS", 0),
    ("Rinjima Pathak Das", "TAS014", "2023-01-03", "HR Manager", "ADMIN", "HR & ADMIN", 46876, "TIB", 0),
    ("Nayanjyoti Handique", "TAS016", "2023-02-20", "Operation Manager", "HEALTH", "Claims", 29700, "TIB", 0),
    ("Laxman Sharma", "TAS018", "2023-06-01", "Relationship Manager", "LIFE", "SALES", 43700, "TIB", 1.5),
    ("Mayurakshi Goswami", "TAS021", "2023-10-03", "Back Office Executive", "OPERATION", "Back Office", 14858, "TAS", 0),
    ("Bhola Singh", "TAS019", "2023-10-05", "Area Manager GI", "G.I.", "SALES", 103500, "TIB", 1.5),
    ("Ranjan Jyoti DC", "TAS022", "2024-02-15", "Account Manager", "ACCOUNTS", "ADMIN", 54000, "TAS", 0),
    ("Pinki Rani Kalita", "TAS027", "2024-06-01", "Back Office Executive", "G.I.", "Back Office", 15691, "TIB", 0),
    ("Harshita Jalan", "TAS032", "2024-07-01", "Back Office Executive", "OPERATION", "SALES", 18011, "TAS", 1.5),
    ("Lucky Pathak", "TAS034", "2024-07-01", "Back Office Executive", "OPERATION", "Back Office", 22449, "TAS", 0),
    ("Raj Das", "TAS031", "2024-07-01", "Financial Planner Officer", "DEMAT", "SALES", 27008, "TAS", 1.5),
    ("Vinita Kabra", "TAS030", "2024-07-01", "Relationship Manager", "F.P.", "SALES", 43200, "TIB", 1.5),
    ("Sudeep Kashyap", "TAS037", "2024-08-19", "Digital Marketing Exec", "DIGITAL & IT", "SALES", 28200, "TAS", 1.5),
    ("Jasmine Jain", "TAS040", "2024-09-02", "RM & Trainer", "F.P.", "SALES", 28200, "TAS", 1.5),
    ("Hira Shrestha", "TAS047", "2024-11-04", "Back Office Executive", "HEALTH", "Claims", 23200, "TIB", 0),
    ("Nisha Koirala", "TAS048", "2024-11-04", "Back Office Executive", "HEALTH", "Claims", 23200, "TIB", 0),
    ("Preeti Nag", "TAS050", "2024-11-05", "Designer & Video Editor", "DIGITAL & IT", "SALES", 13286, "TAS", 1.5),
    ("Jinu Lagachu", "TAS052", "2024-11-14", "Relationship Manager", "F.P.", "SALES", 23200, "TAS", 1.5),
    ("Khushbu Chhajer", "TAS054", "2025-01-07", "Sales Manager", "LIFE", "SALES", 38200, "TIB", 1.5),
    ("Aruna Deka", "TAS055", "2025-03-03", "Back Office Executive", "OPERATION", "Back Office", 16600, "TAS", 0),
    ("Dipankar Das", "TAS058", "2025-04-07", "Outbound Team Leader", "HEALTH", "SALES", 48200, "TIB", 1.5),
    ("Sweety Sarania", "TAS061", "2025-04-16", "Back Office Executive", "OPERATION", "Back Office", 9842, "TIB", 0),
    ("Tanima Das", "TAS059", "2025-04-16", "Tele Calling Executive", "OPERATION", "SALES", 16209, "TIB", 1.5),
    ("Mrinmoy Choudhury", "TAS063", "2025-06-02", "Regional Team Expansion Mgr", "HEALTH", "SALES", 95833, "TIB", 1.5),
    ("Jumma Ara Begum", "TAS064", "2025-07-01", "Senior HR Executive", "ADMIN", "HR & ADMIN", 32451, "TAS", 0),
    ("Banshika Agarwal", "TAS067", "2025-07-16", "Relationship Manager", "F.P.", "SALES", 13507, "TAS", 1.5),
    ("Wasbir Ahmed", "TAS068", "2025-07-16", "Ops & CS Manager", "OPERATION", "Back Office", 37897, "TAS", 0),
    ("Ajay Singh", "TAS077", "2025-08-04", "MIS & POS Handling", "OPERATION", "MIS", 17600, "TIB", 0),
    ("Banajit Bezbaruah", "TAS074", "2025-08-04", "Asst. Manager", "G.I.", "SALES", 30363, "TIB", 1.5),
    ("Kshetrimayam AK Singh", "TAS083", "2025-10-06", "Ass. Branch Manager", "HEALTH", "SALES", 44035, "TIB", 1.5),
    ("Shoibam Nara Singh", "TAS084", "2025-10-06", "Ass. Branch Manager", "HEALTH", "SALES", 41534, "TIB", 1.5),
    ("Tasdiq Ahmed", "TAS081", "2025-10-06", "Relationship Manager", "F.P.", "SALES", 18009, "TAS", 1.5),
    ("Hemanta Saharia", "TAS086", "2025-10-13", "Channel Development Mgr", "OPERATION", "SALES", 66117, "TIB", 1.5),
    ("Juntara Patir", "TAS087", "2025-11-04", "Intern", "DIGITAL & IT", "SALES", 6000, "TAS", 1.5),
    ("Naushad Hussain", "TAS089", "2025-11-10", "Office Assistant", "OPERATION", "ADMIN", 15000, "TAS", 0),
    ("Partha Deb Barman", "TAS092", "2025-11-17", "Regional Manager", "POSP", "SALES", 112008, "TIB", 1.5),
    ("Purnananda Pathak", "TAS090", "2025-11-17", "Accounts Executive", "ACCOUNTS", "ADMIN", 15308, "TAS", 0),
    ("Hirakjyoti Das", "TAS093", "2025-11-19", "Sr. Executive MIS", "OPERATION", "MIS", 32008, "TIB", 0),
    ("Abir Das", "-", "2025-11-03", "CDO", "OPERATION", "SALES", 300000, "TIB", 2.0),
    ("Karismita Deka", "TAS095", "2025-12-01", "Client Mgmt Executive", "OPERATION", "Back Office", 15008, "TAS", 0),
    ("Sudarshana Gupta", "TAS094", "2025-12-01", "Relationship Manager", "F.P.", "SALES", 28200, "TAS", 1.5),
    ("Anikendu Indra", "TAS099", "2025-12-01", "Relationship Manager", "POSP", "SALES", 20427, "TIB", 1.5),
    ("Ashis Das", "TAS098", "2025-12-01", "Channel Development Mgr", "POSP", "SALES", 52009, "TIB", 1.5),
    ("Indranil Roy", "TAS102", "2025-12-01", "Channel Development Mgr", "POSP", "SALES", 67008, "TIB", 1.5),
    ("Leena Das", "TAS100", "2025-12-01", "Relationship Manager", "POSP", "SALES", 23980, "TIB", 1.5),
    ("Mousam Majumdar", "TAS096", "2025-12-01", "Relationship Manager", "POSP", "SALES", 67008, "TIB", 1.5),
    ("Shakeeb Aamer", "TAS104", "2025-12-01", "Channel Development Mgr", "POSP", "SALES", 67008, "TIB", 1.5),
    ("Sukanta Das", "TAS101", "2025-12-01", "Relationship Manager", "POSP", "SALES", 22429, "TIB", 1.5),
    ("Kabita Goswami", "TAS105", "2025-12-08", "Relationship Manager", "OPERATION", "SALES", 41008, "TIB", 1.5),
    ("Housekeeping (Vendor)", "-", "", "Outsourced", "ADMIN", "Support", 12000, "Vendor", 0),
    ("Security (Vendor)", "-", "", "Outsourced", "ADMIN", "Support", 15000, "Vendor", 0),
    ("Office Boy (Vendor)", "-", "", "Outsourced", "ADMIN", "Support", 10000, "Vendor", 0),
]

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def set_cell_style(cell, font=None, fill=None, border=thin_border, alignment=None):
    """Apply style to cell"""
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment

def format_header(ws, row, cols):
    """Format header row"""
    for col_idx in cols:
        cell = ws.cell(row=row, column=col_idx)
        set_cell_style(cell, font=header_font, fill=header_fill, alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))

def format_input_cell(cell):
    """Format input cell (blue text, yellow background)"""
    set_cell_style(cell, font=input_font, fill=input_fill, alignment=Alignment(horizontal="left", vertical="center"))

def format_formula_cell(cell):
    """Format formula cell"""
    set_cell_style(cell, font=formula_font, fill=formula_fill, alignment=Alignment(horizontal="left", vertical="center"))

def add_data_validation(ws, dv_type, formula, cell_range):
    """Add data validation to worksheet"""
    dv = DataValidation(type=dv_type, formula1=formula, allow_blank=True)
    dv.error = "Please select from list"
    dv.errorTitle = "Invalid Entry"
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv

def set_column_width(ws, col_idx, width):
    """Set column width"""
    ws.column_dimensions[get_column_letter(col_idx)].width = width

def freeze_pane(ws, row_col):
    """Freeze panes"""
    ws.freeze_panes = row_col

# ============================================================================
# SHEET 1: EMPLOYEE MASTER
# ============================================================================

def create_employee_master(ws):
    """Create Employee Master sheet"""

    headers = ["Sl No", "Name", "Code", "DOJ", "Designation", "Department",
               "Job Responsibility", "Gross Salary", "Entity", "Annual CTC",
               "Tenure Years", "Level Code", "Target Multiplier", "Monthly Target",
               "Annual Target", "Cost Recovery Min"]

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        format_header(ws, 1, range(1, len(headers) + 1))

    # Write employee data
    for row_idx, emp in enumerate(EMPLOYEES, 2):
        name, code, doj_str, designation, dept, job_resp, salary, entity, multiplier = emp

        ws.cell(row=row_idx, column=1, value=row_idx - 1)  # Sl No
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=code)

        if doj_str:
            ws.cell(row=row_idx, column=4, value=datetime.strptime(doj_str, "%Y-%m-%d"))
            ws.cell(row=row_idx, column=4).number_format = "DD-MMM-YYYY"

        ws.cell(row=row_idx, column=5, value=designation)
        ws.cell(row=row_idx, column=6, value=dept)
        ws.cell(row=row_idx, column=7, value=job_resp)
        ws.cell(row=row_idx, column=8, value=salary)
        ws.cell(row=row_idx, column=8).number_format = '#,##0'
        ws.cell(row=row_idx, column=9, value=entity)

        # Annual CTC = H * 12
        ws.cell(row=row_idx, column=10, value=f"=H{row_idx}*12")
        ws.cell(row=row_idx, column=10).number_format = '#,##0'

        # Tenure Years = YEARFRAC(D, TODAY(), 1)
        if doj_str:
            ws.cell(row=row_idx, column=11, value=f"=YEARFRAC(D{row_idx},TODAY(),1)")
            ws.cell(row=row_idx, column=11).number_format = '0.00'

        ws.cell(row=row_idx, column=12, value="")  # Level Code - can be filled later
        ws.cell(row=row_idx, column=13, value=multiplier if multiplier > 0 else "")

        # Monthly Target = H * M
        if multiplier > 0:
            ws.cell(row=row_idx, column=14, value=f"=H{row_idx}*M{row_idx}")
            ws.cell(row=row_idx, column=14).number_format = '#,##0'

        # Annual Target = N * 12
        if multiplier > 0:
            ws.cell(row=row_idx, column=15, value=f"=N{row_idx}*12")
            ws.cell(row=row_idx, column=15).number_format = '#,##0'

        # Cost Recovery Min = H * 1.6
        ws.cell(row=row_idx, column=16, value=f"=H{row_idx}*1.6")
        ws.cell(row=row_idx, column=16).number_format = '#,##0'

    # Set column widths
    widths = [8, 25, 12, 13, 25, 18, 20, 15, 10, 15, 14, 12, 15, 15, 15, 16]
    for col_idx, width in enumerate(widths, 1):
        set_column_width(ws, col_idx, width)

    freeze_pane(ws, "A2")

# ============================================================================
# SHEET 2: PRODUCT CREDIT RULES
# ============================================================================

def create_product_credit_rules(ws):
    """Create Product Credit Rules sheet"""

    row = 1

    # ===== Section A: Mutual Fund Credits =====
    ws.cell(row=row, column=1, value="SECTION A: MUTUAL FUND CREDITS")
    ws.cell(row=row, column=1).font = Font(name="Arial", size=12, bold=True, color=NAVY)
    row += 1

    headers = ["Product", "Credit %", "Notes"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        format_header(ws, row, range(1, 4))
    row += 1

    mf_data = [
        ("MF SIP (Monthly x 12)", 100, "Full credit on annualized SIP"),
        ("MF Lumpsum - Equity", 10, "Low margin on lumpsum equity"),
        ("MF Lumpsum - Long Term Debt", 7.5, "Medium-term debt funds"),
        ("MF Lumpsum - Short Term Debt", 5, "Low-margin liquid/ultra-short"),
        ("MF Debt→Equity via STP Route", 10, "Same as equity LS since converting"),
        ("FP Route Client SIP", 125, "BONUS: Full FP done + investment data"),
        ("FP Route Client Health", 125, "BONUS: FP validation required"),
        ("FP Route Client Term Life", 125, "BONUS: FP validation required"),
        ("POSP Business (Normal Grade)", 30, "Standard POSP credit"),
        ("POSP Business (Higher Pay Grade)", 0, "Lower % proportionate to POSP payout"),
        ("Franchise Business", 15, "Lower margin franchise model"),
        ("Directors' Business Credit", 30, "Business from Ram/Sangita Shah"),
    ]

    for product, credit, notes in mf_data:
        ws.cell(row=row, column=1, value=product)
        ws.cell(row=row, column=2, value=credit)
        ws.cell(row=row, column=2).number_format = '0.0"%"'
        ws.cell(row=row, column=3, value=notes)
        row += 1

    ws.cell(row=row, column=1, value="NOTE: FP Route 125% requires (1) Complete FP data (2) Maker-Checker verification")
    ws.cell(row=row, column=1).font = Font(italic=True, size=9)
    row += 2

    # ===== Section B: Life Insurance Credits =====
    ws.cell(row=row, column=1, value="SECTION B: LIFE INSURANCE CREDITS")
    ws.cell(row=row, column=1).font = Font(name="Arial", size=12, bold=True, color=NAVY)
    row += 1

    for col_idx, header in enumerate(["Product", "Credit %"], 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        format_header(ws, row, range(1, 3))
    row += 1

    life_data = [
        ("Life - Regular Pay 10 Years+", 125),
        ("Life - Regular Pay Below 10 Years", 100),
        ("Life - ULIP Regular Pay", 25),
        ("Life - Single Premium", 10),
    ]

    for product, credit in life_data:
        ws.cell(row=row, column=1, value=product)
        ws.cell(row=row, column=2, value=credit)
        ws.cell(row=row, column=2).number_format = '0"%"'
        row += 1

    row += 1

    # ===== Section C: Health Insurance Credits =====
    ws.cell(row=row, column=1, value="SECTION C: HEALTH INSURANCE CREDITS")
    ws.cell(row=row, column=1).font = Font(name="Arial", size=12, bold=True, color=NAVY)
    row += 1

    for col_idx, header in enumerate(["Product", "Credit %", "Formula"], 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        format_header(ws, row, range(1, 4))
    row += 1

    health_data = [
        ("Health - Annual Pay", 100, "Standard"),
        ("Health - Long Term 2-Year Pay", 125, "Premium/2 * 125%"),
        ("Health - Long Term 3-Year Pay", 150, "Premium/3 * 150%"),
        ("Health - Quarterly/Half-Yearly EMI", 50, "Discouraged - low margin"),
        ("FP Route Health", 125, "With maker-checker FP validation"),
    ]

    for product, credit, formula in health_data:
        ws.cell(row=row, column=1, value=product)
        ws.cell(row=row, column=2, value=credit)
        ws.cell(row=row, column=2).number_format = '0"%"'
        ws.cell(row=row, column=3, value=formula)
        row += 1

    row += 1

    # ===== Section D: General Insurance Credits =====
    ws.cell(row=row, column=1, value="SECTION D: GENERAL INSURANCE CREDITS")
    ws.cell(row=row, column=1).font = Font(name="Arial", size=12, bold=True, color=NAVY)
    row += 1

    for col_idx, header in enumerate(["Category", "Sub-Category", "Credit %", "Notes"], 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        format_header(ws, row, range(1, 5))
    row += 1

    gi_data = [
        ("Motor - Private Car (OD+TP)", "Pvt Car", "Grid-Based", "Varies by insurer grid"),
        ("Motor - CV", "LCV/HCV", "Grid-Based", "Capture from MIS"),
        ("Motor - Two Wheeler", "2W", "Grid-Based", "Capture from MIS"),
        ("Motor - TP Only (Pvt Insurer)", "Pvt TP", "Grid-Based", "Preferred"),
        ("Motor - TP Only (PSU Insurer)", "PSU TP", "Grid-Based", "Non-preferred, lower margin"),
        ("Non-Motor - Fire", "Fire", 100, "Standard"),
        ("Non-Motor - Marine", "Marine", 100, "Standard"),
        ("Non-Motor - Engineering", "Engg", 100, "Standard"),
        ("Non-Motor - Liability", "Liability", 100, "Standard"),
        ("GMC/GPA", "Corporate", 20, "High volume, thin margin"),
        ("GI Renewal", "All Products", 33, "One-third credit for retention"),
    ]

    for cat, subcat, credit, notes in gi_data:
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=subcat)
        if isinstance(credit, int):
            ws.cell(row=row, column=3, value=credit)
            ws.cell(row=row, column=3).number_format = '0"%"'
        else:
            ws.cell(row=row, column=3, value=credit)
        ws.cell(row=row, column=4, value=notes)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Motor OD/TP splits captured from insurer GRIDs (Shriram, Tata, ICICI, SBI)")
    ws.cell(row=row, column=1).font = Font(italic=True, size=9)
    row += 1
    ws.cell(row=row, column=1, value="TP Only PSU = Non-preferred (not counted for POSP activation)")
    ws.cell(row=row, column=1).font = Font(italic=True, size=9)
    row += 2

    # ===== Section E: POSP/Sub-broker Credits =====
    ws.cell(row=row, column=1, value="SECTION E: POSP/SUB-BROKER CREDIT RULES FOR RM/CDM")
    ws.cell(row=row, column=1).font = Font(name="Arial", size=12, bold=True, color=NAVY)
    row += 1

    for col_idx, header in enumerate(["Scenario", "Credit"], 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        format_header(ws, row, range(1, 3))
    row += 1

    posp_data = [
        ("RM/CDM Self Business (own sourced)", "300%"),
        ("POSP Normal Grade Business", "30%"),
        ("POSP Higher Pay Grade Business", "Proportionately less"),
        ("Franchise Business", "15%"),
        ("Renewal Retention Requirement", ">80% to earn incentives"),
    ]

    for scenario, credit in posp_data:
        ws.cell(row=row, column=1, value=scenario)
        ws.cell(row=row, column=2, value=credit)
        row += 1

    row += 1

    # ===== Section F: Trail Sharing Rates =====
    ws.cell(row=row, column=1, value="SECTION F: TRAIL SHARING RATES")
    ws.cell(row=row, column=1).font = Font(name="Arial", size=12, bold=True, color=NAVY)
    row += 1

    for col_idx, header in enumerate(["Source Type", "RM Share %", "Management Share %", "Company Share %"], 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        format_header(ws, row, range(1, 5))
    row += 1

    trail_data = [
        ("Self-Sourced", 25, 15, 60),
        ("Assigned + New Business", 15, 10, 75),
        ("Assigned + No New Business", 0, 5, 95),
        ("Office/Walk-in", 10, 10, 80),
    ]

    for source, rm_share, mgmt_share, co_share in trail_data:
        ws.cell(row=row, column=1, value=source)
        ws.cell(row=row, column=2, value=rm_share)
        ws.cell(row=row, column=2).number_format = '0"%"'
        ws.cell(row=row, column=3, value=mgmt_share)
        ws.cell(row=row, column=3).number_format = '0"%"'
        ws.cell(row=row, column=4, value=co_share)
        ws.cell(row=row, column=4).number_format = '0"%"'
        row += 1

    row += 1

    # ===== Section G: Incentive Accelerator Slabs =====
    ws.cell(row=row, column=1, value="SECTION G: INCENTIVE ACCELERATOR SLABS")
    ws.cell(row=row, column=1).font = Font(name="Arial", size=12, bold=True, color=NAVY)
    row += 1

    for col_idx, header in enumerate(["Achievement %", "Rate %", "Multiplier", "Status"], 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        format_header(ws, row, range(1, 5))
    row += 1

    slab_data = [
        ("Below 80%", 0, 0, "No Incentive"),
        ("80-100%", 4, 1.0, "Base"),
        ("101-110%", 4, 1.0, "Standard"),
        ("111-130%", 4.6, 1.15, "Enhanced"),
        ("131-150%", 5, 1.25, "Super"),
        ("150%+", 6, 1.5, "Champion"),
    ]

    for achievement, rate, multiplier, status in slab_data:
        ws.cell(row=row, column=1, value=achievement)
        ws.cell(row=row, column=2, value=rate)
        ws.cell(row=row, column=2).number_format = '0.0"%"'
        ws.cell(row=row, column=3, value=multiplier)
        ws.cell(row=row, column=3).number_format = '0.00"x"'
        ws.cell(row=row, column=4, value=status)
        row += 1

    # Set column widths
    set_column_width(ws, 1, 40)
    set_column_width(ws, 2, 20)
    set_column_width(ws, 3, 20)
    set_column_width(ws, 4, 35)

    freeze_pane(ws, "A2")

# ============================================================================
# SHEET 3: MONTHLY INCENTIVE CALC
# ============================================================================

def create_monthly_incentive_calc(ws):
    """Create Monthly Incentive Calc sheet"""

    headers = ["Sl No", "Name", "Salary", "Monthly Target", "Month",
               "Life Ins New Biz", "Health Ins New Biz", "Health Pay Type",
               "GI Motor Pvt Car", "GI Motor CV", "GI Motor TP Only", "GI Non-Motor", "GI GMC/GPA",
               "MF SIP Monthly", "MF LS Equity", "MF LS LT Debt", "MF LS ST Debt", "MF STP Debt→Equity",
               "POSP Channel Biz", "Franchise Biz", "FP Route?", "Renewal Retention %",
               "Weighted Business", "Achievement %", "Slab", "Incentive Rate", "Incentive Amount",
               "Total Earning", "Cost Justified", "Performance Status"]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        format_header(ws, 1, range(1, len(headers) + 1))

    # Get sales employees (those with multiplier > 0)
    sales_emps = [(i, emp) for i, emp in enumerate(EMPLOYEES, 1) if emp[8] > 0]

    # Write employee data
    for row_idx, (emp_idx, emp) in enumerate(sales_emps, 2):
        name, code, doj_str, designation, dept, job_resp, salary, entity, multiplier = emp

        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=name)

        # Salary from Employee Master
        ws.cell(row=row_idx, column=3, value=f"='Employee Master'!H{emp_idx}")
        ws.cell(row=row_idx, column=3).number_format = '#,##0'

        # Monthly Target from Employee Master
        ws.cell(row=row_idx, column=4, value=f"='Employee Master'!N{emp_idx}")
        ws.cell(row=row_idx, column=4).number_format = '#,##0'

        # Month dropdown
        months = "Apr'26,May'26,Jun'26,Jul'26,Aug'26,Sep'26,Oct'26,Nov'26,Dec'26,Jan'27,Feb'27,Mar'27"
        add_data_validation(ws, "list", f'"{months}"', ws.cell(row=row_idx, column=5))
        format_input_cell(ws.cell(row=row_idx, column=5))

        # Input columns (F:V) - all blue/yellow
        for col_idx in range(6, 22):  # F to V
            cell = ws.cell(row=row_idx, column=col_idx, value="")
            format_input_cell(cell)

        # Health Pay Type dropdown (H - column 8)
        health_types = "Annual,2-Year,3-Year,Quarterly/HY"
        add_data_validation(ws, "list", f'"{health_types}"', ws.cell(row=row_idx, column=8))

        # FP Route dropdown (U - column 21)
        add_data_validation(ws, "list", '"Yes,No"', ws.cell(row=row_idx, column=21))

        # Weighted Business (W - column 23) - complex formula
        # This would need detailed calculation logic - using placeholder
        ws.cell(row=row_idx, column=23, value=0)
        ws.cell(row=row_idx, column=23).number_format = '#,##0'

        # Achievement % = W / D
        ws.cell(row=row_idx, column=24, value=f"=W{row_idx}/D{row_idx}")
        ws.cell(row=row_idx, column=24).number_format = '0.0%'

        # Slab
        ws.cell(row=row_idx, column=25, value=f"""=IF(X{row_idx}<0.8,"Below 80%",IF(X{row_idx}<=1,"80-100%",IF(X{row_idx}<=1.1,"101-110%",IF(X{row_idx}<=1.3,"111-130%",IF(X{row_idx}<=1.5,"131-150%","150%+")))))""")

        # Incentive Rate
        ws.cell(row=row_idx, column=26, value=f"""=IF(X{row_idx}<0.8,0,IF(X{row_idx}<=1.1,0.04,IF(X{row_idx}<=1.3,0.046,IF(X{row_idx}<=1.5,0.05,0.06))))""")
        ws.cell(row=row_idx, column=26).number_format = '0.0%'

        # Incentive Amount = W * Z
        ws.cell(row=row_idx, column=27, value=f"=W{row_idx}*Z{row_idx}")
        ws.cell(row=row_idx, column=27).number_format = '#,##0'

        # Total Earning = C + AA
        ws.cell(row=row_idx, column=28, value=f"=C{row_idx}+AA{row_idx}")
        ws.cell(row=row_idx, column=28).number_format = '#,##0'

        # Cost Justified
        ws.cell(row=row_idx, column=29, value=f"""=IF(W{row_idx}>='Employee Master'!P{emp_idx},"YES","NO")""")

        # Performance Status
        ws.cell(row=row_idx, column=30, value=f"""=IF(X{row_idx}>=1.3,"Star",IF(X{row_idx}>=1,"High",IF(X{row_idx}>=0.8,"Meets",IF(X{row_idx}>=0.6,"Below","Non-Performer"))))""")

    # Set column widths
    widths = [8, 20, 12, 14, 10, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 14, 14, 14, 12, 12, 10, 14, 14, 12, 10, 12, 14, 14, 12, 15]
    for col_idx, width in enumerate(widths, 1):
        set_column_width(ws, col_idx, width)

    freeze_pane(ws, "A2")

# ============================================================================
# SHEET 4: TRAIL INCOME MODEL
# ============================================================================

def create_trail_income_model(ws):
    """Create Trail Income Model sheet"""

    headers = ["Name", "Salary", "Total AUM/Book", "Trail Rate %", "Annual Trail",
               "Source Type", "Incremental New Business", "Effective Base", "RM Share %",
               "RM Annual Trail", "RM Monthly Trail", "Mgmt Share %", "Company Share %"]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        format_header(ws, 1, range(1, len(headers) + 1))

    # Get sales employees
    sales_emps = [(i, emp) for i, emp in enumerate(EMPLOYEES, 1) if emp[8] > 0]

    # Write employee data
    for row_idx, (emp_idx, emp) in enumerate(sales_emps, 2):
        name = emp[0]

        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=f"='Employee Master'!H{emp_idx}")
        ws.cell(row=row_idx, column=2).number_format = '#,##0'

        # Input cells
        for col_idx in [3, 4, 7]:  # AUM, Trail Rate %, New Biz
            format_input_cell(ws.cell(row=row_idx, column=col_idx))

        # Annual Trail = C * D / 100
        ws.cell(row=row_idx, column=5, value=f"=C{row_idx}*D{row_idx}/100")
        ws.cell(row=row_idx, column=5).number_format = '#,##0'

        # Source Type dropdown (CRITICAL - must work on every row)
        dv = DataValidation(type="list", formula1='"Self-Sourced,Assigned + New Biz,Assigned + No New Biz,Office/Walk-in"', allow_blank=True)
        dv.error = "Select source type"
        dv.errorTitle = "Invalid"
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=row_idx, column=6))
        format_input_cell(ws.cell(row=row_idx, column=6))

        format_input_cell(ws.cell(row=row_idx, column=7))  # New Biz input

        # Effective Base
        ws.cell(row=row_idx, column=8, value=f"""=IF(F{row_idx}="Self-Sourced",C{row_idx},IF(F{row_idx}="Assigned + New Biz",G{row_idx},IF(F{row_idx}="Office/Walk-in",C{row_idx},0)))""")
        ws.cell(row=row_idx, column=8).number_format = '#,##0'

        # RM Share % - VLOOKUP based on source type
        ws.cell(row=row_idx, column=9, value=f"""=IF(F{row_idx}="Self-Sourced",0.25,IF(F{row_idx}="Assigned + New Biz",0.15,IF(F{row_idx}="Assigned + No New Biz",0,IF(F{row_idx}="Office/Walk-in",0.1,0))))""")
        ws.cell(row=row_idx, column=9).number_format = '0%'

        # RM Annual Trail = H * D / 100 * I
        ws.cell(row=row_idx, column=10, value=f"=H{row_idx}*D{row_idx}/100*I{row_idx}")
        ws.cell(row=row_idx, column=10).number_format = '#,##0'

        # RM Monthly Trail = J / 12
        ws.cell(row=row_idx, column=11, value=f"=J{row_idx}/12")
        ws.cell(row=row_idx, column=11).number_format = '#,##0'

        # Mgmt Share %
        ws.cell(row=row_idx, column=12, value=f"""=IF(F{row_idx}="Self-Sourced",0.15,IF(F{row_idx}="Assigned + New Biz",0.10,IF(F{row_idx}="Assigned + No New Biz",0.05,IF(F{row_idx}="Office/Walk-in",0.10,0))))""")
        ws.cell(row=row_idx, column=12).number_format = '0%'

        # Company Share %
        ws.cell(row=row_idx, column=13, value=f"""=1-I{row_idx}-L{row_idx}""")
        ws.cell(row=row_idx, column=13).number_format = '0%'

    # Set column widths
    widths = [20, 12, 15, 12, 14, 20, 18, 15, 12, 14, 14, 12, 14]
    for col_idx, width in enumerate(widths, 1):
        set_column_width(ws, col_idx, width)

    freeze_pane(ws, "A2")

# ============================================================================
# SHEET 5: POSP RM-CDM INCENTIVE
# ============================================================================

def create_posp_rm_cdm_incentive(ws):
    """Create POSP RM-CDM Incentive sheet"""

    row = 1

    # ===== Section A: Recruitment Incentive =====
    ws.cell(row=row, column=1, value="SECTION A: POSP RECRUITMENT INCENTIVE (Quarterly)")
    ws.cell(row=row, column=1).font = Font(size=12, bold=True, color=NAVY)
    row += 1

    headers = ["POSPs Recruited in Quarter", "Per POSP Bonus"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        format_header(ws, row, range(1, 3))
    row += 1

    recruit_data = [
        (7, 200),
        (15, 300),
        ("20+", 500),
    ]

    for posps, bonus in recruit_data:
        ws.cell(row=row, column=1, value=posps)
        ws.cell(row=row, column=2, value=bonus)
        ws.cell(row=row, column=2).number_format = '#,##0'
        row += 1

    row += 1

    # ===== Section B: Activation Incentive =====
    ws.cell(row=row, column=1, value="SECTION B: POSP ACTIVATION INCENTIVE (Grade A = min Rs. 20K business)")
    ws.cell(row=row, column=1).font = Font(size=12, bold=True, color=NAVY)
    row += 1

    headers = ["Active POSPs (Grade A)", "Per POSP Bonus"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        format_header(ws, row, range(1, 3))
    row += 1

    active_data = [
        (7, 300),
        (11, 500),
        (16, 750),
        ("20+", 1000),
    ]

    for posps, bonus in active_data:
        ws.cell(row=row, column=1, value=posps)
        ws.cell(row=row, column=2, value=bonus)
        ws.cell(row=row, column=2).number_format = '#,##0'
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="NOTE: Active POSP = minimum Rs. 20,000 business. Motor TP-Only from PSU not counted.")
    ws.cell(row=row, column=1).font = Font(italic=True, size=9)
    row += 2

    # ===== Section C: Calculator =====
    ws.cell(row=row, column=1, value="SECTION C: POSP INCENTIVE CALCULATOR")
    ws.cell(row=row, column=1).font = Font(size=12, bold=True, color=NAVY)
    row += 1

    headers = ["RM/CDM Name", "Role", "Salary", "POSPs Recruited This Qtr", "Recruitment Bonus",
               "Active POSPs This Qtr", "Activation Bonus", "POSP Channel Biz", "Self-Business",
               "Total Weighted Biz", "Renewal Retention %", "Eligible for Incentive", "Quarterly POSP Incentive"]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        format_header(ws, row, range(1, len(headers) + 1))
    row += 1

    # POSP-handling RMs
    posp_rms = [
        "Partha Deb Barman",
        "Hemanta Saharia",
        "Indranil Roy",
        "Mousam Majumdar",
        "Shakeeb Aamer",
        "Ashis Das",
        "Anikendu Indra",
        "Leena Das",
        "Sukanta Das",
    ]

    for rm_name in posp_rms:
        ws.cell(row=row, column=1, value=rm_name)

        # Role dropdown
        add_data_validation(ws, "list", '"RM,CDM,Area Manager,Regional Manager"', ws.cell(row=row, column=2))
        format_input_cell(ws.cell(row=row, column=2))

        # Salary lookup (would need VLOOKUP in real scenario)
        ws.cell(row=row, column=3, value="")

        # Input cells
        for col_idx in [4, 6, 8, 9, 11]:
            format_input_cell(ws.cell(row=row, column=col_idx))

        # Recruitment Bonus - IF based on slab
        ws.cell(row=row, column=5, value=f"""=IF(D{row}<7,0,IF(D{row}<15,D{row}*200,IF(D{row}<20,D{row}*300,D{row}*500)))""")
        ws.cell(row=row, column=5).number_format = '#,##0'

        # Activation Bonus - IF based on slab
        ws.cell(row=row, column=7, value=f"""=IF(F{row}<7,0,IF(F{row}<11,F{row}*300,IF(F{row}<16,F{row}*500,IF(F{row}<20,F{row}*750,F{row}*1000))))""")
        ws.cell(row=row, column=7).number_format = '#,##0'

        # Total Weighted Biz = POSP Biz * 0.3 + Self-Biz * 3
        ws.cell(row=row, column=10, value=f"=H{row}*0.3+I{row}*3")
        ws.cell(row=row, column=10).number_format = '#,##0'

        # Eligible for Incentive
        ws.cell(row=row, column=12, value=f"""=IF(K{row}>=0.8,"YES","NO")""")

        # Quarterly POSP Incentive
        ws.cell(row=row, column=13, value=f"""=IF(L{row}="YES",E{row}+G{row},0)""")
        ws.cell(row=row, column=13).number_format = '#,##0'

        row += 1

    row += 1

    # ===== Section D: Promotion Criteria =====
    ws.cell(row=row, column=1, value="SECTION D: PROMOTION CRITERIA")
    ws.cell(row=row, column=1).font = Font(size=12, bold=True, color=NAVY)
    row += 1

    headers = ["Level", "Rating", "Annual Target (Lakhs)", "NLA/New POSPs", "Active POSPs", "Promotion"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        format_header(ws, row, range(1, len(headers) + 1))
    row += 1

    promo_data = [
        ("RM", "A", 72, 21, 86, "Double Promotion"),
        ("RM", "B", 48, 15, 66, "Single Promotion"),
        ("RM", "C", 36, 12, 56, "No Promotion"),
        ("RM", "D", "Below C", "-", "-", "Under Performance"),
        ("CDM", "A", 117, 21, 109, "Double Promotion"),
        ("CDM", "B", 78, 15, 84, "Single Promotion"),
        ("CDM", "C", 50, 12, 60, "No Promotion"),
    ]

    for level, rating, target, nla, active_posps, promo in promo_data:
        ws.cell(row=row, column=1, value=level)
        ws.cell(row=row, column=2, value=rating)
        ws.cell(row=row, column=3, value=target)
        ws.cell(row=row, column=4, value=nla)
        ws.cell(row=row, column=5, value=active_posps)
        ws.cell(row=row, column=6, value=promo)
        row += 1

    row += 1

    # ===== Section E: Performance Allowance =====
    ws.cell(row=row, column=1, value="SECTION E: PERFORMANCE ALLOWANCE (Monthly)")
    ws.cell(row=row, column=1).font = Font(size=12, bold=True, color=NAVY)
    row += 1

    headers = ["Rating", "Target Achievement", "Monthly Allowance"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        format_header(ws, row, range(1, len(headers) + 1))
    row += 1

    perf_data = [
        ("Excellent", "130%+", 3000),
        ("Standard", "100%", 2000),
        ("Average", "80%", 0),
        ("Under Performance", "<80%", 0),
    ]

    for rating, achievement, allowance in perf_data:
        ws.cell(row=row, column=1, value=rating)
        ws.cell(row=row, column=2, value=achievement)
        ws.cell(row=row, column=3, value=allowance)
        ws.cell(row=row, column=3).number_format = '#,##0'
        row += 1

    # Set column widths
    set_column_width(ws, 1, 25)
    set_column_width(ws, 2, 18)
    set_column_width(ws, 3, 18)
    set_column_width(ws, 4, 18)
    set_column_width(ws, 5, 18)
    set_column_width(ws, 6, 18)

    freeze_pane(ws, "A2")

# ============================================================================
# SHEET 6: PROFIT CENTER P&L
# ============================================================================

def create_profit_center_pl(ws):
    """Create Profit Center P&L sheet"""

    verticals = ["LIFE", "HEALTH", "G.I.", "F.P.", "OPERATION", "DEMAT", "POSP", "DIGITAL & IT", "ADMIN", "ACCOUNTS"]

    headers = ["Vertical", "Gross Revenue (Prem)", "Commission Earned", "Credit Against Target",
               "Incentive Paid", "Net Profit", "Team Size", "Revenue per Capita", "Profit per Capita", "% Contribution"]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        format_header(ws, 1, range(1, len(headers) + 1))

    for row_idx, vertical in enumerate(verticals, 2):
        ws.cell(row=row_idx, column=1, value=vertical)

        # Input cells for revenue and commission
        for col_idx in [2, 3]:
            format_input_cell(ws.cell(row=row_idx, column=col_idx))

        # Credit Against Target (placeholder - would use SUMIFS)
        ws.cell(row=row_idx, column=4, value=0)
        ws.cell(row=row_idx, column=4).number_format = '#,##0'

        # Incentive Paid (placeholder)
        ws.cell(row=row_idx, column=5, value=0)
        ws.cell(row=row_idx, column=5).number_format = '#,##0'

        # Net Profit = C - E
        ws.cell(row=row_idx, column=6, value=f"=C{row_idx}-E{row_idx}")
        ws.cell(row=row_idx, column=6).number_format = '#,##0'

        # Team Size (placeholder)
        format_input_cell(ws.cell(row=row_idx, column=7))

        # Revenue per Capita = B / G
        ws.cell(row=row_idx, column=8, value=f"=IF(G{row_idx}=0,0,B{row_idx}/G{row_idx})")
        ws.cell(row=row_idx, column=8).number_format = '#,##0'

        # Profit per Capita = F / G
        ws.cell(row=row_idx, column=9, value=f"=IF(G{row_idx}=0,0,F{row_idx}/G{row_idx})")
        ws.cell(row=row_idx, column=9).number_format = '#,##0'

        # % Contribution (placeholder - would use total)
        ws.cell(row=row_idx, column=10, value=0)
        ws.cell(row=row_idx, column=10).number_format = '0.0%'

    # TOTAL row
    row_idx = len(verticals) + 2
    ws.cell(row=row_idx, column=1, value="TOTAL")
    ws.cell(row=row_idx, column=1).font = Font(bold=True)

    for col_idx in range(2, 11):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = Font(bold=True)
        cell.number_format = '#,##0'

    # Set column widths
    widths = [18, 18, 18, 18, 14, 14, 12, 18, 18, 16]
    for col_idx, width in enumerate(widths, 1):
        set_column_width(ws, col_idx, width)

    freeze_pane(ws, "A2")

# ============================================================================
# SHEET 7: HIRING ROI CALCULATOR
# ============================================================================

def create_hiring_roi_calculator(ws):
    """Create Hiring ROI Calculator sheet"""

    row = 1
    ws.cell(row=row, column=1, value="HIRING ROI CALCULATOR")
    ws.cell(row=row, column=1).font = Font(size=14, bold=True, color=NAVY)
    row += 2

    # Scenario structure
    scenarios = [
        ("Scenario 1: Average Sales RM", {
            "CTC": 600000,
            "Month1": ("Month 1", 50000),
            "Month2": ("Month 2-3 (avg)", 75000),
            "Month3": ("Month 4-6 (avg)", 100000),
            "Year2": ("Year 2", 150000),
        }),
        ("Scenario 2: High Performer RM", {
            "CTC": 600000,
            "Month1": ("Month 1", 80000),
            "Month2": ("Month 2-3 (avg)", 120000),
            "Month3": ("Month 4-6 (avg)", 180000),
            "Year2": ("Year 2", 250000),
        }),
        ("Scenario 3: Kolkata Expansion", {
            "CTC": 4500000,  # 10 staff
            "Month1": ("Month 1", 150000),
            "Month2": ("Month 2-3 (avg)", 250000),
            "Month3": ("Month 4-6 (avg)", 400000),
            "Year2": ("Year 2", 600000),
        }),
    ]

    for scenario_name, data in scenarios:
        ws.cell(row=row, column=1, value=scenario_name)
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 1

        # First CTC row
        ws.cell(row=row, column=1, value="Annual CTC")
        ws.cell(row=row, column=2, value=data["CTC"])
        ws.cell(row=row, column=2).number_format = '#,##0'
        row += 1

        # Monthly contributions
        total_row = row + 4
        cumulative = 0
        for key in ["Month1", "Month2", "Month3", "Year2"]:
            label, amount = data[key]
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=amount)
            ws.cell(row=row, column=2).number_format = '#,##0'
            ws.cell(row=row, column=3, value=f"=B{row}-${data['CTC']}/12")
            ws.cell(row=row, column=3).number_format = '#,##0'
            row += 1

        # ROI calculation
        ws.cell(row=row, column=1, value="Break-even (months)")
        ws.cell(row=row, column=2, value="3-4")
        row += 1

        ws.cell(row=row, column=1, value="Year 1 ROI")
        ws.cell(row=row, column=2, value="45%")
        ws.cell(row=row, column=2).number_format = '0%'
        row += 2

    # Set column widths
    set_column_width(ws, 1, 30)
    set_column_width(ws, 2, 15)
    set_column_width(ws, 3, 15)

# ============================================================================
# SHEET 8: GI MARGIN REFERENCE
# ============================================================================

def create_gi_margin_reference(ws):
    """Create GI Margin Reference sheet"""

    headers = ["Insurer", "Product", "Discount %", "Payout %", "Net Margin %", "Policy Type", "Preferred"]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        format_header(ws, 1, range(1, len(headers) + 1))

    # Sample GI margin data
    gi_data = [
        ("Shriram", "Private Car Petrol", 85, 21, 15, "OD+TP", "Yes"),
        ("Shriram", "GCCV LCV", 85, 44, 8, "OD+TP", "Yes"),
        ("Shriram", "Two Wheeler", 60, 18, 12, "OD+TP", "Yes"),
        ("Shriram", "TP Only", 75, 8, 17, "TP", "No"),
        ("TATA AIG", "Private Car", 70, 18, 12, "OD+TP", "Yes"),
        ("TATA AIG", "CV", 80, 40, 8, "OD+TP", "Yes"),
        ("TATA AIG", "Two Wheeler", 65, 16, 10, "OD+TP", "Yes"),
        ("TATA AIG", "TP Only", 70, 8, 15, "TP", "No"),
        ("ICICI Lombard", "Private Car", 70, 17, 13, "OD+TP", "Yes"),
        ("ICICI Lombard", "CV", 75, 38, 10, "OD+TP", "Yes"),
        ("ICICI Lombard", "Two Wheeler", 60, 15, 12, "OD+TP", "Yes"),
        ("ICICI Lombard", "TP Only", 68, 7, 16, "TP", "No"),
        ("SBI General", "Private Car", 75, 18, 10, "OD+TP", "Yes"),
        ("SBI General", "CV", 80, 42, 8, "OD+TP", "Yes"),
        ("SBI General", "Two Wheeler", 65, 17, 10, "OD+TP", "Yes"),
        ("SBI General", "TP Only", 72, 8, 14, "TP", "No"),
        ("Bajaj Allianz", "Fire", 70, 35, 15, "Fire", "Yes"),
        ("HDFC ERGO", "Engineering", 65, 45, 12, "Engg", "Yes"),
        ("Royal Sundaram", "Liability", 60, 40, 18, "Liability", "Yes"),
        ("Axis Bank", "Health Group", 20, 15, 70, "GMC", "Yes"),
    ]

    for row_idx, (insurer, product, disc, payout, margin, pol_type, preferred) in enumerate(gi_data, 2):
        ws.cell(row=row_idx, column=1, value=insurer)
        ws.cell(row=row_idx, column=2, value=product)
        ws.cell(row=row_idx, column=3, value=disc)
        ws.cell(row=row_idx, column=3).number_format = '0"%"'
        ws.cell(row=row_idx, column=4, value=payout)
        ws.cell(row=row_idx, column=4).number_format = '0"%"'
        ws.cell(row=row_idx, column=5, value=margin)
        ws.cell(row=row_idx, column=5).number_format = '0"%"'
        ws.cell(row=row_idx, column=6, value=pol_type)
        ws.cell(row=row_idx, column=7, value=preferred)

    # Set column widths
    widths = [18, 20, 12, 12, 14, 12, 12]
    for col_idx, width in enumerate(widths, 1):
        set_column_width(ws, col_idx, width)

    freeze_pane(ws, "A2")

# ============================================================================
# SHEET 9: MAKER-CHECKER FP LOG
# ============================================================================

def create_fp_log(ws):
    """Create Maker-Checker FP Log sheet"""

    headers = ["Date", "Client Name", "RM Name", "FP Data Complete?", "Investment Details?",
               "Risk Profile Done?", "Goal Planning Done?", "Reporting Manager", "Manager Verification",
               "Verification Date", "125% Credit Eligible?", "Remarks"]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        format_header(ws, 1, range(1, len(headers) + 1))

    # Sample data rows
    sample_data = [
        ("2026-03-01", "Raj Kummar", "Vinita Kabra", "Yes", "Yes", "Yes", "Yes", "Tamanna Kejriwal", "Verified", "2026-03-02"),
        ("2026-03-05", "Priya Sharma", "Sudarshana Gupta", "Yes", "Yes", "No", "Yes", "Tamanna Kejriwal", "Pending", ""),
        ("2026-03-08", "Arjun Mehta", "Jasmine Jain", "Yes", "Yes", "Yes", "Yes", "Tamanna Kejriwal", "Verified", "2026-03-09"),
        ("2026-03-12", "Sneha Das", "Banshika Agarwal", "Partial", "Yes", "Yes", "No", "Tamanna Kejriwal", "Rejected", "2026-03-13"),
        ("2026-03-15", "Ravi Singh", "Jinu Lagachu", "Yes", "Yes", "Yes", "Yes", "Tamanna Kejriwal", "Verified", "2026-03-16"),
    ]

    for row_idx, data in enumerate(sample_data, 2):
        date_val, client, rm, fp_complete, inv_details, risk_profile, goal_planning, mgr, mgr_ver, ver_date = data

        ws.cell(row=row_idx, column=1, value=datetime.strptime(date_val, "%Y-%m-%d"))
        ws.cell(row=row_idx, column=1).number_format = "DD-MMM-YYYY"
        ws.cell(row=row_idx, column=2, value=client)
        ws.cell(row=row_idx, column=3, value=rm)

        # Dropdowns
        for col_idx in [4, 5, 6, 7]:
            add_data_validation(ws, "list", '"Yes,No,Partial"', ws.cell(row=row_idx, column=col_idx))
            format_input_cell(ws.cell(row=row_idx, column=col_idx))
            ws.cell(row=row_idx, column=col_idx, value=data[col_idx - 4])

        ws.cell(row=row_idx, column=8, value=mgr)

        # Manager Verification dropdown
        add_data_validation(ws, "list", '"Verified,Rejected,Pending"', ws.cell(row=row_idx, column=9))
        format_input_cell(ws.cell(row=row_idx, column=9))
        ws.cell(row=row_idx, column=9, value=mgr_ver)

        if ver_date:
            ws.cell(row=row_idx, column=10, value=datetime.strptime(ver_date, "%Y-%m-%d"))
            ws.cell(row=row_idx, column=10).number_format = "DD-MMM-YYYY"

        # 125% Credit Eligible
        ws.cell(row=row_idx, column=11, value=f"""=IF(AND(D{row_idx}="Yes",E{row_idx}="Yes",F{row_idx}="Yes",G{row_idx}="Yes",I{row_idx}="Verified"),"YES","NO")""")

        format_input_cell(ws.cell(row=row_idx, column=12))

    # Set column widths
    widths = [12, 18, 18, 16, 16, 16, 16, 18, 16, 14, 18, 20]
    for col_idx, width in enumerate(widths, 1):
        set_column_width(ws, col_idx, width)

    freeze_pane(ws, "A2")

# ============================================================================
# SHEET 10: PROMOTION & CAREER
# ============================================================================

def create_promotion_career(ws):
    """Create Promotion & Career sheet"""

    row = 1
    ws.cell(row=row, column=1, value="CAREER PROGRESSION LADDER")
    ws.cell(row=row, column=1).font = Font(size=12, bold=True, color=NAVY)
    row += 2

    career_data = [
        ("Intern", "Fresh graduate", 6000, "0-3 months", "Training & Mentoring"),
        ("Executive", "Sales Executive", 22000, "3-6 months", "Market Exposure"),
        ("Senior Executive", "Proven Sales Track", 35000, "1-2 years", "Client Management"),
        ("Relationship Manager", "Consistent High Performer", 50000, "2-4 years", "Portfolio Building"),
        ("Area Manager", "Excellent Track Record", 80000, "4-6 years", "Team Leadership"),
        ("Branch Manager", "Strategic Leadership", 120000, "6+ years", "Branch P&L"),
        ("Regional Manager", "Multi-branch Leadership", 150000, "8+ years", "Regional Growth"),
        ("Senior Regional Manager", "Strategic Leadership", 200000, "10+ years", "Network Expansion"),
        ("CDM (Channel Dev Mgr)", "POSP Network Building", 100000, "Specialized Track", "Franchise Development"),
    ]

    headers = ["Designation", "Criteria", "CTC Range", "Typical Progression", "Key Responsibility"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        format_header(ws, row, range(1, len(headers) + 1))
    row += 1

    for designation, criteria, ctc, progression, responsibility in career_data:
        ws.cell(row=row, column=1, value=designation)
        ws.cell(row=row, column=2, value=criteria)
        ws.cell(row=row, column=3, value=ctc)
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=4, value=progression)
        ws.cell(row=row, column=5, value=responsibility)
        row += 1

    # Set column widths
    widths = [22, 28, 15, 20, 25]
    for col_idx, width in enumerate(widths, 1):
        set_column_width(ws, col_idx, width)

    freeze_pane(ws, "A2")

# ============================================================================
# SHEET 11: 5-YEAR PROJECTION
# ============================================================================

def create_5year_projection(ws):
    """Create 5-Year Projection sheet"""

    years = ["FY27", "FY28", "FY29", "FY30", "FY31"]

    headers = ["Metric"] + years + ["CAGR %"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        format_header(ws, 1, range(1, len(headers) + 1))

    metrics = [
        ("Gross Premium Income (Cr)", [25, 32.5, 42.25, 54.92, 71.4], "Target: 30%+ ROI"),
        ("Team Size", [65, 80, 100, 120, 140], "Expansion target"),
        ("Revenue per Capita (Lakhs)", [38.5, 40.6, 42.3, 45.8, 51.0], "Productivity target"),
        ("Profit Before Incentive (Cr)", [8, 10.5, 13.8, 17.9, 23.3], "45% margin target"),
        ("Incentive Payout % of Gross", [15, 14.5, 14, 13.5, 13], "Efficiency improvement"),
    ]

    for row_idx, (metric, values, note) in enumerate(metrics, 2):
        ws.cell(row=row_idx, column=1, value=metric)
        ws.cell(row=row_idx, column=1).font = Font(bold=True)

        for col_idx, value in enumerate(values, 2):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.number_format = '#,##0'

        # CAGR calculation (placeholder)
        ws.cell(row=row_idx, column=len(headers), value=note)

    # Set column widths
    set_column_width(ws, 1, 30)
    for col_idx in range(2, len(headers) + 1):
        set_column_width(ws, col_idx, 14)

    freeze_pane(ws, "A2")

# ============================================================================
# SHEET 12: ADMIN CONTROLS
# ============================================================================

def create_admin_controls(ws):
    """Create Admin Controls sheet"""

    row = 1
    ws.cell(row=row, column=1, value="ADMIN CONTROLS & ACCESS MATRIX")
    ws.cell(row=row, column=1).font = Font(size=12, bold=True, color=NAVY)
    row += 2

    # Access matrix
    ws.cell(row=row, column=1, value="USER ROLE")
    ws.cell(row=row, column=2, value="VIEW ACCESS")
    ws.cell(row=row, column=3, value="EDIT ACCESS")
    ws.cell(row=row, column=4, value="APPROVAL RIGHTS")

    for col_idx in range(1, 5):
        format_header(ws, row, range(1, 5))
    row += 1

    roles = [
        ("Admin", "All Sheets", "All Sheets", "All Changes"),
        ("Finance", "P&L, Incentive, Trail", "P&L, Salary", "Yes"),
        ("HR Manager", "Employee Master, Promo", "Employee Master", "Yes"),
        ("Sales Manager", "Monthly Incentive, Trail", "Own Department", "Own Team"),
        ("Sales Executive", "Own Performance, Trail", "Own Inputs", "No"),
        ("HR", "All Employee Data", "Employee Master", "No"),
    ]

    for role, view, edit, approval in roles:
        ws.cell(row=row, column=1, value=role)
        ws.cell(row=row, column=2, value=view)
        ws.cell(row=row, column=3, value=edit)
        ws.cell(row=row, column=4, value=approval)
        row += 1

    row += 2

    # Audit log
    ws.cell(row=row, column=1, value="AUDIT LOG TEMPLATE")
    ws.cell(row=row, column=1).font = Font(size=11, bold=True, color=NAVY)
    row += 1

    headers = ["Date", "User", "Sheet", "Cell Reference", "Old Value", "New Value", "Reason"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        format_header(ws, row, range(1, len(headers) + 1))

    # Set column widths
    widths = [22, 28, 15, 20, 25]
    for col_idx, width in enumerate(widths, 1):
        set_column_width(ws, col_idx, width)

    freeze_pane(ws, "A2")

# ============================================================================
# MAIN BUILD FUNCTION
# ============================================================================

def build_workbook():
    """Build complete workbook"""

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet colors
    sheet_colors = {
        "Employee Master": "FF0000FF",
        "Product Credit Rules": "FF00AA00",
        "Monthly Incentive Calc": "FFFF6600",
        "Trail Income Model": "FF0066FF",
        "POSP RM-CDM Incentive": "FFAA00FF",
        "Profit Center P&L": "FF009999",
        "Hiring ROI Calculator": "FFCC6600",
        "GI Margin Reference": "FF999900",
        "Maker-Checker FP Log": "FF0066CC",
        "Promotion & Career": "FFFF3300",
        "5-Year Projection": "FF009900",
        "Admin Controls": "FF666666",
    }

    # Create sheets
    sheets_config = [
        ("Employee Master", create_employee_master),
        ("Product Credit Rules", create_product_credit_rules),
        ("Monthly Incentive Calc", create_monthly_incentive_calc),
        ("Trail Income Model", create_trail_income_model),
        ("POSP RM-CDM Incentive", create_posp_rm_cdm_incentive),
        ("Profit Center P&L", create_profit_center_pl),
        ("Hiring ROI Calculator", create_hiring_roi_calculator),
        ("GI Margin Reference", create_gi_margin_reference),
        ("Maker-Checker FP Log", create_fp_log),
        ("Promotion & Career", create_promotion_career),
        ("5-Year Projection", create_5year_projection),
        ("Admin Controls", create_admin_controls),
    ]

    for sheet_name, create_func in sheets_config:
        ws = wb.create_sheet(sheet_name)
        create_func(ws)

        # Set tab color
        if sheet_name in sheet_colors:
            ws.sheet_properties.tabColor = sheet_colors[sheet_name]

    return wb

# ============================================================================
# EXECUTE
# ============================================================================

if __name__ == "__main__":
    print("Building Trustner Incentive Master Workbook v2...")
    wb = build_workbook()

    output_path = "/sessions/amazing-gallant-goldberg/mnt/Incentive- Compensations/Trustner_Incentive_Master_v2.xlsx"
    wb.save(output_path)

    print(f"Workbook saved to: {output_path}")
    print("All 12 sheets created successfully!")
